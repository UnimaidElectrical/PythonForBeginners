#!/usr/bin/env python

# export data sheets from xlsx to csv

from openpyxl import load_workbook
import csv
from os import sys

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file,read_only=True,data_only=True)
    all_worksheets = workbook.sheetnames
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets
get_all_sheets('excel_to_convert.xlsx')


def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file, data_only=True)
    for worksheet_name in sheets:
        print("Export " + worksheet_name + " ...")

        try:
            worksheet = workbook.get_sheet_by_name(worksheet_name)
            print(worksheet)
        except KeyError:
            print("Could not find " + worksheet_name)
            sys.exit(1)

        your_csv_file = open(''.join([worksheet_name, '.csv']), 'w')
        wx = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                lrow.append(cell.value)
            wx.writerow(lrow)
        print(" ... done")
    your_csv_file.close()
#get_all_sheets('excel_to_convert.xlsx')
#csv_from_excel('excel_to_convert.xlsx',get_all_sheets('excel_to_convert.xlsx'))


if not 2 <= len(sys.argv) <= 3:
	print("Call with " + sys.argv[0] + " <xlxs file> [comma separated list of sheets to export]")
	sys.exit(1)
else:
    sheets = []
    if len(sys.argv) == 3:
        sheets = list(sys.argv[2].split(','))
    else:
        sheets = get_all_sheets(sys.argv[1])
    assert(sheets != None and len(sheets) > 0)
    csv_from_excel(sys.argv[1], sheets)







'''
from openpyxl import load_workbook
import csv
from os import sys

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    all_worksheets = workbook.get_sheet_names
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    #print(f'{sheets} this are the files')
    return sheets



def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file, data_only=True)
    for worksheet_name in sheets:
        print("Export " + worksheet_name + " ...")

        try:
            worksheet = workbook.get_sheet_by_name(worksheet_name)
        except KeyError:
            print("Could not find " + worksheet_name)
            sys.exit(1)

        your_csv_file = open(''.join([worksheet_name, '.csv']), 'wb')
        wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                lrow.append(cell.value)
            wr.writerow(lrow)
        print(" ... done")
    your_csv_file.close()



#get_all_sheets(excel_to_convert.xslx)


if not 2 <= len(sys.argv) <= 3:
	print("Call with " + sys.argv[0] + " <xlxs file> [comma separated list of sheets to export]")
	sys.exit(1)
else:
    sheets = []
    if len(sys.argv) == 3:
        sheets = list(sys.argv[2].split(','))
    else:
        sheets = get_all_sheets(sys.argv[1])
    assert(sheets != None and len(sheets) > 0)
    csv_from_excel(sys.argv[1], sheets)
'''





'''
use_iterators
read_only




d=[3,4,5,6,7,8,9]
for i in range(len(d)):
    print(i)



x=2.0000
n=10
c= n**x
print(str(c))






#!/usr/bin/env python

# export data sheets from xlsx to csv

from openpyxl import load_workbook
import csv
from os import sys





import sys
print( "This is the name of the script: ", sys.argv[0])
print ("Number of arguments: ", len(sys.argv))
print ("The arguments are: " , str(sys.argv))

'''

"""

strings=[i.lower() for i in s if i.isalnum()]




def checking(mad):
    strings = []
    for i in mad:
        if i.isalnum():
            i.lower()

ands=checking("A man")
print(ands)






def checking(mad):
    strings=[i.lower() for i in mad if i.isalnum()]
    return strings

ands=checking("A man")
print(ands)



[<value_when_condition_true> if <condition> else <value_when_condition_false> for value in list_name]

outputlist = []
for y in a:
    if y not in b:
        outputlist.append(y)

"""